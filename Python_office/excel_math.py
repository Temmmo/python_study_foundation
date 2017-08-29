import openpyxl
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.get_active_sheet()
print(sheet['A3'].value)

wbda = openpyxl.load_workbook('writeFormula.xlsx',data_only=True)
sheet = wbda.get_active_sheet()
print(sheet['A3'].value)



