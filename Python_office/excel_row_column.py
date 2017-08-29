#! python
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

wb1 = openpyxl.Workbook()
sheet1 = wb1.get_active_sheet()
sheet1.merge_cells('A1:D3')
sheet1['A1'] = 'Twelve cells merged together'
sheet1.merge_cells('C5:D5')
sheet1['C5'] = 'Two merged cells'
wb1.save('merged.xlsx')


