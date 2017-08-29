#! python3
#  openpyxl --- study
import openpyxl, os
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(type(sheet))
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)
print(anotherSheet['A1'].value)
c=anotherSheet['B1']
print(c.value)
print('Row'+ str(c.row)+',Column '+c.column+' is'+ c.value)
print('Cell'+ c.coordinate +' is' + c.value)

sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1,8,2):
    print(i,sheet.cell(row=i, column=2).value)

#print(sheet.get_highest_row())
#print(sheet.get_highest_column())
