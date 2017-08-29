#! python
#! excel ----study

import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(get_column_letter(sheet.get_highest_column()))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))
print(tuple(sheet['A1':'C3']))

for row in sheet['A1':'C3']:
    for cellObj in row:
        print(cellObj.coordinate,cellObj.value)
    print('-----end of row')

sheet.columns[1]
for cellObj in sheet.columns[1]:
    print(cellObj.value)
